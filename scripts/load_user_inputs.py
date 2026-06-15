#!/usr/bin/env python3
"""Phase UIL Task 1 (B1): user-input loader.

Reads ``production_run/MODEL_INPUTS_TEMPLATE.xlsx`` by **tab name + header**
(openpyxl), validates every user-entered field, and writes a normalised,
schema-versioned ``model_inputs.json`` for the run orchestrator (B3).

Pure I/O + validation -- NO model math. Frozen/governed parameters in the
template (copula df, grouped-t dfs, Sigma) are read back for echo/provenance
only and are never user-settable.

Fail-loud contract: every problem is reported as
``Tab '<tab>', row <n>, field '<field>': <message>`` and all problems found
are listed together before the loader exits non-zero.

Usage:
    python3 scripts/load_user_inputs.py \
        --template production_run/MODEL_INPUTS_TEMPLATE.xlsx \
        --out model_inputs.json
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

SCHEMA_VERSION = "1.0.0"

ALLOWED_PRODUCT_TYPES = ("HKCD_PAR_2026", "HKRB_PAR_2026", "GMMB_EQ_2026")
ALLOWED_GENDERS = ("M", "F")
ALLOWED_SCALES = ("units", "thousands", "millions")
ALLOWED_THOUSANDS = ("comma", "space", "period", "none")

REQUIRED_TABS = ("Currency", "Balance Sheet", "Portfolio", "Assumptions", "Run Settings")

PORTFOLIO_HEADER = (
    "Product type", "Issue age", "Gender", "Term (yrs)",
    "Sum assured", "Annual premium", "Policy count", "Vested bonus",
)


class InputValidationError(ValueError):
    """Raised when the template fails validation; message lists every issue."""

    def __init__(self, errors: List[str]):
        self.errors = list(errors)
        msg = "MODEL_INPUTS_TEMPLATE validation failed with %d issue(s):\n  - %s" % (
            len(self.errors), "\n  - ".join(self.errors))
        super().__init__(msg)


# ----------------------------------------------------------------- helpers
def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _as_str(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _to_float(v: Any) -> Optional[float]:
    if _is_blank(v):
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _to_int(v: Any) -> Optional[int]:
    f = _to_float(v)
    if f is None or f != int(f):
        return None
    return int(f)


def _err(tab: str, row: Any, field: str, message: str) -> str:
    return "Tab '%s', row %s, field '%s': %s" % (tab, row, field, message)


def _find_field_rows(ws) -> Dict[str, Tuple[int, Any]]:
    """Map first-column label -> (row_number, value-in-column-B)."""
    out: Dict[str, Tuple[int, Any]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and not _is_blank(row[0]):
            label = _as_str(row[0])
            if label not in out:
                out[label] = (i, row[1] if len(row) > 1 else None)
    return out


def _lookup(fields: Dict[str, Tuple[int, Any]], prefix: str) -> Optional[Tuple[int, Any]]:
    """Find a field row whose label starts with ``prefix`` (template labels are stable)."""
    for label, rv in fields.items():
        if label.startswith(prefix):
            return rv
    return None


def _require(fields, prefix, tab, field_name, errors) -> Optional[Tuple[int, Any]]:
    hit = _lookup(fields, prefix)
    if hit is None:
        errors.append(_err(tab, "-", field_name, "required field row not found (label starting with '%s')" % prefix))
    return hit


# ----------------------------------------------------------------- tab parsers
def parse_currency(ws, errors: List[str]) -> Dict[str, Any]:
    tab = "Currency"
    f = _find_field_rows(ws)
    out: Dict[str, Any] = {}

    hit = _require(f, "Reporting currency code", tab, "Reporting currency code", errors)
    if hit:
        code = _as_str(hit[1]).upper()
        if len(code) != 3 or not code.isalpha():
            errors.append(_err(tab, hit[0], "Reporting currency code", "must be a 3-letter ISO 4217 code, got %r" % code))
        out["code"] = code

    hit = _require(f, "Reporting currency symbol", tab, "Reporting currency symbol", errors)
    if hit:
        sym = _as_str(hit[1])
        if not sym:
            errors.append(_err(tab, hit[0], "Reporting currency symbol", "must not be empty"))
        out["symbol"] = sym

    hit = _require(f, "Decimal places", tab, "Decimal places", errors)
    if hit:
        dec = _to_int(hit[1])
        if dec is None or dec < 0 or dec > 6:
            errors.append(_err(tab, hit[0], "Decimal places", "must be an integer in [0, 6], got %r" % hit[1]))
        else:
            out["decimals"] = dec

    hit = _require(f, "Amount scale", tab, "Amount scale", errors)
    if hit:
        scale = _as_str(hit[1]).lower()
        if scale not in ALLOWED_SCALES:
            errors.append(_err(tab, hit[0], "Amount scale", "must be one of %s, got %r" % (list(ALLOWED_SCALES), hit[1])))
        out["scale"] = scale

    hit = _require(f, "Thousands separator", tab, "Thousands separator", errors)
    if hit:
        sep = _as_str(hit[1]).lower()
        if sep not in ALLOWED_THOUSANDS:
            errors.append(_err(tab, hit[0], "Thousands separator", "must be one of %s, got %r" % (list(ALLOWED_THOUSANDS), hit[1])))
        out["thousands"] = sep

    hit = _require(f, "Calibration market label", tab, "Calibration market label", errors)
    if hit:
        lbl = _as_str(hit[1])
        if not lbl:
            errors.append(_err(tab, hit[0], "Calibration market label", "must not be empty"))
        out["market_label"] = lbl

    hit = _require(f, "Valuation date", tab, "Valuation date", errors)
    if hit:
        raw = hit[1]
        if isinstance(raw, _dt.datetime):
            out["valuation_date"] = raw.date().isoformat()
        elif isinstance(raw, _dt.date):
            out["valuation_date"] = raw.isoformat()
        else:
            s = _as_str(raw)
            try:
                out["valuation_date"] = _dt.date.fromisoformat(s).isoformat()
            except ValueError:
                errors.append(_err(tab, hit[0], "Valuation date", "must be YYYY-MM-DD, got %r" % raw))
    return out


def parse_balance_sheet(ws, errors: List[str]) -> Dict[str, Any]:
    tab = "Balance Sheet"
    rows = list(ws.iter_rows(values_only=True))
    out: Dict[str, Any] = {"assets": []}

    header_row = None
    for i, row in enumerate(rows, 1):
        if row and _as_str(row[0]) == "Asset class":
            header_row = i
            break
    if header_row is None:
        errors.append(_err(tab, "-", "Asset class", "asset table header row not found"))
        return out

    total_mv = 0.0
    illiquid_mv = 0.0
    for i in range(header_row, len(rows)):
        row = rows[i]
        rno = i + 1
        label = _as_str(row[0]) if row else ""
        if not label or label.startswith("Total backing"):
            break
        mv = _to_float(row[1] if len(row) > 1 else None)
        if mv is None or mv < 0:
            errors.append(_err(tab, rno, "Market value", "must be a number >= 0, got %r" % (row[1] if len(row) > 1 else None)))
            continue
        illiq_raw = _as_str(row[2] if len(row) > 2 else "").lower()
        if illiq_raw not in ("yes", "no"):
            errors.append(_err(tab, rno, "Illiquid?", "must be 'Yes' or 'No', got %r" % (row[2] if len(row) > 2 else None)))
            continue
        illiquid = illiq_raw == "yes"
        out["assets"].append({"asset_class": label, "market_value": mv, "illiquid": illiquid})
        total_mv += mv
        if illiquid:
            illiquid_mv += mv

    if not out["assets"]:
        errors.append(_err(tab, header_row, "Asset class", "no asset rows found under the header"))
    if total_mv <= 0:
        errors.append(_err(tab, header_row, "Market value", "total backing asset market value must be positive (got %s)" % total_mv))

    out["backing_asset_mv"] = total_mv
    out["illiquid_mv"] = illiquid_mv
    out["illiquid_share"] = (illiquid_mv / total_mv) if total_mv > 0 else None
    if out["illiquid_share"] is not None and not (0.0 <= out["illiquid_share"] <= 1.0):
        errors.append(_err(tab, "-", "Illiquid share (derived)", "derived share %r outside [0, 1]" % out["illiquid_share"]))

    f = _find_field_rows(ws)
    hit = _lookup(f, "Total backing asset market value")
    if hit is not None:
        stated = _to_float(hit[1])
        if stated is not None and total_mv > 0 and abs(stated - total_mv) > max(1e-6 * total_mv, 1e-9):
            errors.append(_err(tab, hit[0], "Total backing asset market value",
                               "stated total %s does not match the sum of asset rows %s -- re-save the workbook so derived cells recalculate" % (stated, total_mv)))

    hit = _require(f, "Forced-sale fraction", tab, "Forced-sale fraction (mass-lapse shock)", errors)
    if hit:
        fs = _to_float(hit[1])
        if fs is None or not (0.0 < fs <= 1.0):
            errors.append(_err(tab, hit[0], "Forced-sale fraction (mass-lapse shock)", "must be in (0, 1], got %r" % hit[1]))
        else:
            out["forced_sale_fraction"] = fs

    hit = _require(f, "Best-estimate liability", tab, "Best-estimate liability (reserve)", errors)
    if hit:
        bel = _to_float(hit[1])
        if bel is None or bel <= 0:
            errors.append(_err(tab, hit[0], "Best-estimate liability (reserve)", "must be a positive number, got %r" % hit[1]))
        else:
            out["best_estimate_liability"] = bel

    hit = _require(f, "Equity-guarantee initial index level", tab, "Equity-guarantee initial index level", errors)
    if hit:
        idx = _to_float(hit[1])
        if idx is None or idx <= 0:
            errors.append(_err(tab, hit[0], "Equity-guarantee initial index level", "must be a positive number, got %r" % hit[1]))
        else:
            out["equity_guarantee_initial_index"] = idx
    return out


def parse_portfolio(ws, errors: List[str]) -> List[Dict[str, Any]]:
    tab = "Portfolio"
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows, 1):
        if row and tuple(_as_str(c) for c in row[:8]) == PORTFOLIO_HEADER:
            header_row = i
            break
    if header_row is None:
        errors.append(_err(tab, "-", "header", "expected header %s not found" % (list(PORTFOLIO_HEADER),)))
        return []

    points: List[Dict[str, Any]] = []
    for i in range(header_row, len(rows)):
        row = rows[i]
        rno = i + 1
        if row is None or all(_is_blank(c) for c in row):
            continue
        first = _as_str(row[0])
        # stop at totals/footnote furniture
        if not first:
            if len(row) > 3 and _as_str(row[3]).startswith("Totals"):
                continue
            continue
        if first.startswith("Product types:"):
            break

        product = first
        if product not in ALLOWED_PRODUCT_TYPES:
            errors.append(_err(tab, rno, "Product type", "%r not in allowed set %s" % (product, list(ALLOWED_PRODUCT_TYPES))))
            continue

        def need(idx: int, name: str):
            v = row[idx] if len(row) > idx else None
            if _is_blank(v):
                errors.append(_err(tab, rno, name, "row is incomplete -- value required"))
                return None
            return v

        age = _to_int(need(1, "Issue age"))
        gender = _as_str(need(2, "Gender")).upper() or None
        term = _to_int(need(3, "Term (yrs)"))
        sa = _to_float(need(4, "Sum assured"))
        prem = _to_float(need(5, "Annual premium"))
        count = _to_int(need(6, "Policy count"))
        vb = _to_float(need(7, "Vested bonus"))

        if age is not None and not (0 <= age <= 120):
            errors.append(_err(tab, rno, "Issue age", "must be an integer in [0, 120], got %r" % row[1]))
        if gender is not None and gender not in ALLOWED_GENDERS:
            errors.append(_err(tab, rno, "Gender", "must be one of %s, got %r" % (list(ALLOWED_GENDERS), row[2])))
        if term is not None and term <= 0:
            errors.append(_err(tab, rno, "Term (yrs)", "must be a positive integer, got %r" % row[3]))
        if sa is not None and sa <= 0:
            errors.append(_err(tab, rno, "Sum assured", "must be positive, got %r" % row[4]))
        if prem is not None and prem < 0:
            errors.append(_err(tab, rno, "Annual premium", "must be >= 0, got %r" % row[5]))
        if count is not None and count <= 0:
            errors.append(_err(tab, rno, "Policy count", "must be a positive integer, got %r" % row[6]))
        if vb is not None and vb < 0:
            errors.append(_err(tab, rno, "Vested bonus", "must be >= 0, got %r" % row[7]))

        if None in (age, gender, term, sa, prem, count, vb):
            continue  # incompleteness already reported field-by-field
        points.append({
            "product_type": product, "issue_age": age, "gender": gender,
            "term_years": term, "sum_assured": sa, "annual_premium": prem,
            "policy_count": count, "vested_bonus": vb, "source_row": rno,
        })

    if not points and not errors:
        errors.append(_err(tab, header_row + 1, "Product type", "at least one complete model-point row is required"))
    return points


def parse_assumptions(ws, errors: List[str]) -> Dict[str, Any]:
    tab = "Assumptions"
    f = _find_field_rows(ws)
    out: Dict[str, Any] = {}

    def bounded(prefix, key, lo, hi, lo_open=True, hi_open=False, field_name=None):
        name = field_name or prefix
        hit = _require(f, prefix, tab, name, errors)
        if not hit:
            return
        v = _to_float(hit[1])
        ok = v is not None
        if ok:
            ok = (v > lo if lo_open else v >= lo) and (v < hi if hi_open else v <= hi)
        if not ok:
            ival = "%s%s, %s%s" % ("(" if lo_open else "[", lo, hi, ")" if hi_open else "]")
            errors.append(_err(tab, hit[0], name, "must be in %s, got %r" % (ival, hit[1])))
        else:
            out[key] = v

    bounded("Confidence level (SCR)", "confidence", 0.0, 1.0, lo_open=True, hi_open=True)
    bounded("Management-action relief: si", "relief_sigma", 0.0, 10.0, lo_open=True, hi_open=False,
            field_name="Management-action relief: sigma")
    bounded("Management-action relief: al", "relief_alpha", 0.0, 1.0, lo_open=True, hi_open=False,
            field_name="Management-action relief: alpha")
    bounded("Benefit share (beta_fit)", "benefit_share", 0.0, 1.0, lo_open=True, hi_open=False)

    # governed / frozen read-back (provenance echo only -- never user-settable)
    frozen: Dict[str, Any] = {}
    for prefix, key in (
        ("Copula degrees of freedom", "copula_df_single_t"),
        ("Grouped-t df — NON-FIN block", "grouped_t_df_nonfin"),
        ("Grouped-t df — FIN block", "grouped_t_df_fin"),
    ):
        hit = _lookup(f, prefix)
        if hit is not None:
            frozen[key] = _to_float(hit[1])
    out["governed_frozen_readback"] = frozen
    return out


def parse_run_settings(ws, errors: List[str]) -> Dict[str, Any]:
    tab = "Run Settings"
    f = _find_field_rows(ws)
    out: Dict[str, Any] = {}

    for prefix, key, mini in (
        ("Number of simulations", "n_sim", 1),
        ("Bootstrap replicates", "bootstrap_replicates", 1),
        ("Projection horizon", "horizon_months", 1),
    ):
        hit = _require(f, prefix, tab, prefix, errors)
        if hit:
            v = _to_int(hit[1])
            if v is None or v < mini:
                errors.append(_err(tab, hit[0], prefix, "must be an integer >= %d, got %r" % (mini, hit[1])))
            else:
                out[key] = v

    hit = _require(f, "Random seed", tab, "Random seed", errors)
    if hit:
        v = _to_int(hit[1])
        if v is None:
            errors.append(_err(tab, hit[0], "Random seed", "must be an integer, got %r" % hit[1]))
        else:
            out["seed"] = v

    hit = _require(f, "Output label", tab, "Output label / scenario name", errors)
    if hit:
        lbl = _as_str(hit[1])
        if not lbl:
            errors.append(_err(tab, hit[0], "Output label / scenario name", "must not be empty"))
        else:
            out["output_label"] = lbl
    return out



# ----------------------------------------------------------------- run-controls (no-Excel) validator
def validate_run_controls_dict(payload: Dict[str, Any]) -> List[str]:
    """Validate a GUI-produced run-controls fragment (the ``{currency,
    run_settings}`` subset of model_inputs.json) WITHOUT an Excel template,
    using the SAME rules as the template parsers. Returns the full list of
    issues (empty == valid); the fail-loud message format mirrors the template
    path. This is the loader-side validator the Phase IGUI input+run GUI
    round-trips every payload through before a run is permitted (no openpyxl
    needed). Purely additive: the Excel path above is unchanged.
    """
    errors: List[str] = []
    if not isinstance(payload, dict):
        return ["run controls payload must be a JSON object"]
    cur = payload.get("currency") or {}
    rs = payload.get("run_settings") or {}
    if not isinstance(cur, dict):
        errors.append(_err("Currency", "-", "currency", "must be an object"))
        cur = {}
    if not isinstance(rs, dict):
        errors.append(_err("Run Settings", "-", "run_settings", "must be an object"))
        rs = {}

    # --- currency ---
    code = _as_str(cur.get("code")).upper()
    if len(code) != 3 or not code.isalpha():
        errors.append(_err("Currency", "-", "Reporting currency code",
                           "must be a 3-letter ISO 4217 code, got %r" % cur.get("code")))
    if _is_blank(cur.get("symbol")):
        errors.append(_err("Currency", "-", "Reporting currency symbol", "must not be empty"))
    scale = _as_str(cur.get("scale"))
    if scale not in ALLOWED_SCALES:
        errors.append(_err("Currency", "-", "Amount scale",
                           "must be one of %s, got %r" % (ALLOWED_SCALES, cur.get("scale"))))
    thousands = _as_str(cur.get("thousands"))
    if thousands not in ALLOWED_THOUSANDS:
        errors.append(_err("Currency", "-", "Thousands separator",
                           "must be one of %s, got %r" % (ALLOWED_THOUSANDS, cur.get("thousands"))))
    if _is_blank(cur.get("market_label")):
        errors.append(_err("Currency", "-", "Calibration market label", "must not be empty"))
    vd = _as_str(cur.get("valuation_date"))
    try:
        _dt.date.fromisoformat(vd)
    except ValueError:
        errors.append(_err("Currency", "-", "Valuation date",
                           "must be YYYY-MM-DD, got %r" % cur.get("valuation_date")))

    # --- run settings ---
    for key, field_name, mini in (
        ("n_outer", "Outer scenarios", 1),
        ("n_inner", "Inner paths", 1),
        ("n_sim", "Number of simulations", 1),
        ("bootstrap_replicates", "Bootstrap replicates", 1),
        ("horizon_months", "Projection horizon", 1),
        ("step_months", "Projection step", 1),
    ):
        v = _to_int(rs.get(key))
        if v is None or v < mini:
            errors.append(_err("Run Settings", "-", field_name,
                               "must be an integer >= %d, got %r" % (mini, rs.get(key))))
    horizon = _to_int(rs.get("horizon_months"))
    step = _to_int(rs.get("step_months"))
    if horizon is not None and step is not None and step >= 1 and horizon >= 1:
        if step > horizon:
            errors.append(_err("Run Settings", "-", "Projection step",
                               "must not exceed the projection horizon (%d > %d)" % (step, horizon)))
        elif horizon % step != 0:
            errors.append(_err("Run Settings", "-", "Projection step",
                               "must divide the projection horizon evenly (%d %% %d != 0)" % (horizon, step)))
    if _to_int(rs.get("seed")) is None:
        errors.append(_err("Run Settings", "-", "Random seed",
                           "must be an integer, got %r" % rs.get("seed")))
    if _is_blank(rs.get("output_label")):
        errors.append(_err("Run Settings", "-", "Output label / scenario name", "must not be empty"))
    digest = rs.get("reproducibility_digest")
    if digest is not None and not (
            isinstance(digest, str) and digest.startswith("sha256:") and len(digest) == 71):
        errors.append(_err("Run Settings", "-", "reproducibility_digest",
                           "must be 'sha256:<64 hex>' when present"))
    return errors


# ------------------------------------------------- portfolio (no-Excel) validator
def validate_portfolio_dict(payload: Dict[str, Any]) -> List[str]:
    """Validate a GUI-produced portfolio fragment (the ``{portfolio,
    balance_sheet}`` subset of model_inputs.json) WITHOUT an Excel template,
    using the SAME rules as :func:`parse_portfolio` / :func:`parse_balance_sheet`.

    Returns the full list of issues (empty == valid); the fail-loud message
    format mirrors the template path. This is the loader-side validator the
    Phase IGUI input+run GUI (Task 3, model points + in-force ingest) round-trips
    every payload through before a run is permitted (no openpyxl needed). Purely
    additive: the Excel parsers above are unchanged.
    """
    errors: List[str] = []
    if not isinstance(payload, dict):
        return ["portfolio payload must be a JSON object"]

    # --- portfolio rows ---
    rows = payload.get("portfolio")
    if not isinstance(rows, list):
        errors.append(_err("Portfolio", "-", "portfolio", "must be a JSON array of model-point rows"))
        rows = []
    if isinstance(rows, list) and len(rows) == 0:
        errors.append(_err("Portfolio", "-", "Product type",
                           "at least one complete model-point row is required"))
    n_par = 0
    for i, row in enumerate(rows, 1):
        rno = row.get("source_row", i) if isinstance(row, dict) else i
        if not isinstance(row, dict):
            errors.append(_err("Portfolio", rno, "row", "must be a JSON object"))
            continue
        product = _as_str(row.get("product_type"))
        if product not in ALLOWED_PRODUCT_TYPES:
            errors.append(_err("Portfolio", rno, "Product type",
                               "%r not in allowed set %s" % (product, list(ALLOWED_PRODUCT_TYPES))))
        age = _to_int(row.get("issue_age"))
        gender = _as_str(row.get("gender")).upper() or None
        term = _to_int(row.get("term_years"))
        sa = _to_float(row.get("sum_assured"))
        prem = _to_float(row.get("annual_premium"))
        count = _to_int(row.get("policy_count"))
        vb = _to_float(row.get("vested_bonus"))
        if age is None or not (0 <= age <= 120):
            errors.append(_err("Portfolio", rno, "Issue age", "must be an integer in [0, 120], got %r" % row.get("issue_age")))
        if gender is None or gender not in ALLOWED_GENDERS:
            errors.append(_err("Portfolio", rno, "Gender", "must be one of %s, got %r" % (list(ALLOWED_GENDERS), row.get("gender"))))
        if term is None or term <= 0:
            errors.append(_err("Portfolio", rno, "Term (yrs)", "must be a positive integer, got %r" % row.get("term_years")))
        if sa is None or sa <= 0:
            errors.append(_err("Portfolio", rno, "Sum assured", "must be positive, got %r" % row.get("sum_assured")))
        if prem is None or prem < 0:
            errors.append(_err("Portfolio", rno, "Annual premium", "must be >= 0, got %r" % row.get("annual_premium")))
        if count is None or count <= 0:
            errors.append(_err("Portfolio", rno, "Policy count", "must be a positive integer, got %r" % row.get("policy_count")))
        if vb is None or vb < 0:
            errors.append(_err("Portfolio", rno, "Vested bonus", "must be >= 0, got %r" % row.get("vested_bonus")))
        if product == "HKCD_PAR_2026" and vb is not None and vb > 0:
            errors.append(_err("Portfolio", rno, "Vested bonus",
                               "cash-dividend PAR cannot carry a vested reversionary bonus, got %r" % vb))
        if product in ("HKCD_PAR_2026", "HKRB_PAR_2026"):
            n_par += 1
    if isinstance(rows, list) and len(rows) > 0 and n_par == 0:
        errors.append(_err("Portfolio", "-", "Product type",
                           "at least one PAR model point is required (only GMMB rows supplied)"))

    # --- balance sheet (optional block; validated when present) ---
    bs = payload.get("balance_sheet")
    if bs is not None:
        if not isinstance(bs, dict):
            errors.append(_err("Balance Sheet", "-", "balance_sheet", "must be a JSON object"))
        else:
            assets = bs.get("assets")
            if not isinstance(assets, list) or not assets:
                errors.append(_err("Balance Sheet", "-", "Asset class", "no asset rows found"))
                assets = assets if isinstance(assets, list) else []
            total_mv = 0.0
            for j, a in enumerate(assets, 1):
                if not isinstance(a, dict):
                    errors.append(_err("Balance Sheet", j, "asset row", "must be a JSON object"))
                    continue
                mv = _to_float(a.get("market_value"))
                if mv is None or mv < 0:
                    errors.append(_err("Balance Sheet", j, "Market value", "must be a number >= 0, got %r" % a.get("market_value")))
                else:
                    total_mv += mv
                if not isinstance(a.get("illiquid"), bool):
                    errors.append(_err("Balance Sheet", j, "Illiquid?", "must be a boolean, got %r" % a.get("illiquid")))
            if total_mv <= 0:
                errors.append(_err("Balance Sheet", "-", "Market value", "total backing asset market value must be positive (got %s)" % total_mv))
            stated = bs.get("stated_total_backing_asset_mv")
            if stated is not None:
                stated_f = _to_float(stated)
                if stated_f is not None and total_mv > 0 and abs(stated_f - total_mv) > max(1e-6 * total_mv, 1e-9):
                    errors.append(_err("Balance Sheet", "-", "Total backing asset market value",
                                       "stated total %s does not match the sum of asset rows %s" % (stated_f, total_mv)))
            fs = _to_float(bs.get("forced_sale_fraction"))
            if fs is None or not (0.0 < fs <= 1.0):
                errors.append(_err("Balance Sheet", "-", "Forced-sale fraction (mass-lapse shock)", "must be in (0, 1], got %r" % bs.get("forced_sale_fraction")))
            bel = _to_float(bs.get("best_estimate_liability"))
            if bel is None or bel <= 0:
                errors.append(_err("Balance Sheet", "-", "Best-estimate liability (reserve)", "must be a positive number, got %r" % bs.get("best_estimate_liability")))
            idx = _to_float(bs.get("equity_guarantee_initial_index"))
            if idx is None or idx <= 0:
                errors.append(_err("Balance Sheet", "-", "Equity-guarantee initial index level", "must be a positive number, got %r" % bs.get("equity_guarantee_initial_index")))
    return errors



# ----------------------------------------------------------------- main API
def load_user_inputs(template_path: str) -> Dict[str, Any]:
    """Parse + validate the template; return the normalised inputs dict.

    Raises InputValidationError (with the full issue list) on any problem.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    if not os.path.exists(template_path):
        raise InputValidationError(["template file not found: %s" % template_path])

    wb = openpyxl.load_workbook(template_path, data_only=True)
    errors: List[str] = []
    missing = [t for t in REQUIRED_TABS if t not in wb.sheetnames]
    if missing:
        raise InputValidationError(
            ["Tab '%s', row -, field '-': required tab missing from workbook" % t for t in missing])

    currency = parse_currency(wb["Currency"], errors)
    balance_sheet = parse_balance_sheet(wb["Balance Sheet"], errors)
    portfolio = parse_portfolio(wb["Portfolio"], errors)
    assumptions = parse_assumptions(wb["Assumptions"], errors)
    run_settings = parse_run_settings(wb["Run Settings"], errors)

    if errors:
        raise InputValidationError(errors)

    total_sa = sum(p["sum_assured"] * p["policy_count"] for p in portfolio)
    policy_count = sum(p["policy_count"] for p in portfolio)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _dt.datetime.now(_dt.timezone.utc).isoformat(),
        "source_template": os.path.basename(template_path),
        "currency": currency,
        "balance_sheet": balance_sheet,
        "portfolio": portfolio,
        "assumptions": assumptions,
        "run_settings": run_settings,
        "totals": {
            "backing_asset_mv": balance_sheet.get("backing_asset_mv"),
            "total_sum_assured": total_sa,
            "policy_count": policy_count,
            "model_point_rows": len(portfolio),
        },
    }


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Validate the model-inputs template and write model_inputs.json")
    ap.add_argument("--template", default="production_run/MODEL_INPUTS_TEMPLATE.xlsx")
    ap.add_argument("--out", default="model_inputs.json")
    args = ap.parse_args(argv)

    try:
        inputs = load_user_inputs(args.template)
    except InputValidationError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(inputs, fh, indent=1)
    # re-parse what we wrote: fail loudly rather than hand a corrupt file downstream
    with open(args.out, "r", encoding="utf-8") as fh:
        json.load(fh)

    cur = inputs["currency"]; tot = inputs["totals"]
    print("model_inputs.json written: %s (schema %s)" % (args.out, SCHEMA_VERSION))
    print("  currency            : %s (%s)  scale=%s" % (cur["code"], cur["symbol"], cur["scale"]))
    print("  total asset MV      : %s" % format(tot["backing_asset_mv"], ",.2f"))
    print("  total sum assured   : %s  (SA x policy count)" % format(tot["total_sum_assured"], ",.2f"))
    print("  policy count        : %s in %d model-point row(s)" % (format(tot["policy_count"], ","), tot["model_point_rows"]))
    print("  run                 : n_sim=%d seed=%d label=%s" % (
        inputs["run_settings"]["n_sim"], inputs["run_settings"]["seed"], inputs["run_settings"]["output_label"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
